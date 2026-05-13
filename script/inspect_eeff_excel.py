import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def main() -> None:
    file_path = Path("public/docs/EEFF GESTIÓN S.xlsx")
    xls = pd.ExcelFile(file_path)
    wb = load_workbook(file_path, data_only=False)
    summary = {"file": str(file_path), "sheets": []}

    for sheet in xls.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)
        ws = wb[sheet]
        non_empty_rows = int(df.notna().any(axis=1).sum())
        non_empty_cols = int(df.notna().any(axis=0).sum())
        preview = df.iloc[:12, :12].fillna("").astype(str).values.tolist()
        formula_count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        summary["sheets"].append(
            {
                "name": sheet,
                "shape": [int(df.shape[0]), int(df.shape[1])],
                "non_empty_rows": non_empty_rows,
                "non_empty_cols": non_empty_cols,
                "formula_cells": formula_count,
                "preview_12x12": preview,
            }
        )

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
