import json
import re
from pathlib import Path

from openpyxl import load_workbook

MESES = [
    "ENERO",
    "FEBRERO",
    "MARZO",
    "ABRIL",
    "MAYO",
    "JUNIO",
    "JULIO",
    "AGOSTO",
    "SEPTIEMBRE",
    "OCTUBRE",
    "NOVIEMBRE",
    "DICIEMBRE",
]


def to_number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip()
    if not txt or txt in {"-", "—", "$", "�"}:
        return 0.0
    txt = txt.replace("$", "").replace(" ", "")
    txt = txt.replace(".", "").replace(",", ".")
    txt = re.sub(r"[^0-9\-.]", "", txt)
    if txt in {"", "-", ".", "-."}:
        return 0.0
    try:
        return float(txt)
    except ValueError:
        return 0.0


def parse_sheet(ws):
    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        b = str(ws.cell(row=r, column=2).value or "").upper()
        if "RESULTADO OPERACIONAL" in b:
            header_row = r
            break
    if header_row is None:
        return None

    month_cols = []
    for c in range(4, ws.max_column + 1):
        label = str(ws.cell(row=header_row, column=c).value or "").strip().upper()
        label = re.sub(r"\s+", " ", label)
        if any(m in label for m in MESES):
            month_name = next(m for m in MESES if m in label)
            month_cols.append((month_name, c))
            if len(month_cols) == 12:
                break
        elif month_cols:
            break
    if not month_cols:
        return None

    cuentas = []
    for r in range(header_row + 1, ws.max_row + 1):
        code_raw = str(ws.cell(row=r, column=1).value or "").strip()
        desc = str(ws.cell(row=r, column=2).value or "").strip()
        if not code_raw or not re.match(r"^\d+(?:-\d+)+$", code_raw):
            continue
        monthly = {}
        non_zero = False
        for month_name, col in month_cols:
            v = to_number(ws.cell(row=r, column=col).value)
            monthly[month_name] = v
            if abs(v) > 0:
                non_zero = True
        if not non_zero:
            continue
        cuentas.append({"codigo": code_raw, "nombre": desc, "mensual": monthly})
    return {"meses": [m for m, _ in month_cols], "cuentas": cuentas}


def main():
    source = Path("public/docs/EEFF GESTIÓN S.xlsx")
    target = Path("src/assets/data/FIDELMIRA/eeff_excel.json")
    wb = load_workbook(source, data_only=True)

    payload = {"origen": str(source), "anios": {}}
    for sheet in wb.sheetnames:
        m = re.search(r"(20\d{2})", sheet)
        if not m:
            continue
        parsed = parse_sheet(wb[sheet])
        if not parsed:
            continue
        payload["anios"][m.group(1)] = parsed

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {target}")


if __name__ == "__main__":
    main()
